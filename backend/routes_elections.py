import os
import logging
from typing import Dict, List, Any, Optional
from fastapi import APIRouter, HTTPException, Request, status, UploadFile, File, Depends
from pydantic import BaseModel
import httpx

logger = logging.getLogger("sdps.elections")

# Supabase REST settings
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
if SUPABASE_URL.endswith("/rest/v1"):
    SUPABASE_URL = SUPABASE_URL[:-8]
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

headers = {
    "apikey": SUPABASE_KEY or "",
    "Authorization": f"Bearer {SUPABASE_KEY}" if SUPABASE_KEY else "",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

elections_router = APIRouter(prefix="/api/elections", tags=["elections"])

# Models
class VoteCastPayload(BaseModel):
    admission_no: str
    selections: Dict[str, str]

class SettingUpdatePayload(BaseModel):
    value: str

class PostCreatePayload(BaseModel):
    key: str
    title: str
    order_index: int

class CandidateCreatePayload(BaseModel):
    name: str
    post_key: str
    symbol: str
    photo: Optional[str] = ""
    symbol_image: Optional[str] = ""

async def check_supabase():
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise HTTPException(
            status_code=503,
            detail="Supabase service is not configured on the backend."
        )

# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@elections_router.get("/settings")
async def get_settings():
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.get(f"{SUPABASE_URL}/rest/v1/election_settings", headers=headers)
        if r.status_code != 200:
            logger.error(f"Supabase settings fetch failed: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to fetch settings.")
        
        data = r.json()
        settings = {}
        for d in data:
            settings[d["key"]] = d["value"]
        return settings

@elections_router.get("/stats")
async def get_stats():
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_voters?select=admission_no",
            headers={**headers, "Prefer": "count=exact", "Range": "0-0"}
        )
        count_header = r.headers.get("content-range")
        vCount = 0
        if count_header and "/" in count_header:
            vCount = int(count_header.split("/")[-1])
        return {"voters_count": vCount}

@elections_router.get("/voters/{admission_no}")
async def get_voter(admission_no: str):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_voters?admission_no=eq.{admission_no}",
            headers=headers
        )
        if r.status_code != 200:
            logger.error(f"Voter fetch error: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to search voter.")
        
        voters = r.json()
        if not voters:
            raise HTTPException(status_code=404, detail="Voter not found in roster.")
        
        v = voters[0]
        return {
            "admission_no": v["admission_no"],
            "name": v["name"],
            "role": v["role"],
            "has_voted": v["already_voted"],
            "class_name": v.get("class_name"),
            "father_name": v.get("father_name")
        }

@elections_router.get("/bootstrap")
async def bootstrap():
    await check_supabase()
    async with httpx.AsyncClient() as client:
        # Fetch posts
        r_posts = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_posts?order=order_index.asc",
            headers=headers
        )
        # Fetch candidates
        r_cands = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_candidates",
            headers=headers
        )

        if r_posts.status_code != 200 or r_cands.status_code != 200:
            logger.error(f"Bootstrap fetch failed: posts={r_posts.text}, candidates={r_cands.text}")
            raise HTTPException(status_code=500, detail="Failed to load ballot.")

        posts = [{"key": p["key"], "title": p["title"], "order": p["order_index"]} for p in r_posts.json()]
        candidates = [{
            "id": c["id"],
            "post": c["post_key"],
            "name": c["name"],
            "photo": c.get("photo", ""),
            "symbol": c.get("symbol", ""),
            "symbol_image": c.get("symbol_image", ""),
            "adjustment": c.get("adjustment", 0)
        } for c in r_cands.json()]

        return {
            "posts": posts,
            "candidates": candidates
        }

@elections_router.post("/vote")
async def cast_vote(payload: VoteCastPayload):
    await check_supabase()
    adm = payload.admission_no.strip()
    selections = payload.selections

    async with httpx.AsyncClient() as client:
        # 1. Check election settings
        r_settings = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_settings?key=eq.election_open",
            headers=headers
        )
        if r_settings.status_code != 200 or not r_settings.json():
            raise HTTPException(status_code=403, detail="Voting is currently closed.")
        
        if r_settings.json()[0]["value"] != "true":
            raise HTTPException(status_code=403, detail="Voting is currently closed.")

        # 2. Check voter status
        r_voter = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_voters?admission_no=eq.{adm}",
            headers=headers
        )
        if r_voter.status_code != 200 or not r_voter.json():
            raise HTTPException(status_code=404, detail="Voter not registered.")
        
        voter = r_voter.json()[0]
        if voter["already_voted"]:
            raise HTTPException(status_code=400, detail="You have already casted your vote.")

        # 3. Insert selections into election_votes
        r_vote = await client.post(
            f"{SUPABASE_URL}/rest/v1/election_votes",
            json={"voter_admission_no": adm, "selections": selections},
            headers=headers
        )
        if r_vote.status_code not in [200, 201]:
            logger.error(f"Insert vote failed: {r_vote.text}")
            raise HTTPException(status_code=500, detail="Failed to cast vote.")

        # 4. Set voter status to already voted
        r_update = await client.patch(
            f"{SUPABASE_URL}/rest/v1/election_voters?admission_no=eq.{adm}",
            json={"already_voted": True},
            headers=headers
        )
        if r_update.status_code not in [200, 204]:
            logger.error(f"Voter update failed: {r_update.text}")
            # Rollback vote insert
            await client.delete(
                f"{SUPABASE_URL}/rest/v1/election_votes?voter_admission_no=eq.{adm}",
                headers=headers
            )
            raise HTTPException(status_code=500, detail="Voter registration confirmation failed.")

        return {"success": True, "message": "Vote cast successfully!"}

# ─────────────────────────────────────────────────────────────────────────────
# ADMIN ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@elections_router.post("/settings/{key}")
async def update_settings(key: str, payload: SettingUpdatePayload):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        # Upsert setting
        r = await client.post(
            f"{SUPABASE_URL}/rest/v1/election_settings",
            json={"key": key, "value": payload.value},
            headers={**headers, "Prefer": "resolution=merge-duplicates"}
        )
        if r.status_code not in [200, 201, 204]:
            logger.error(f"Failed to update setting {key}: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to save settings.")
        return {"success": True}

@elections_router.post("/posts")
async def create_post(payload: PostCreatePayload):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.post(
            f"{SUPABASE_URL}/rest/v1/election_posts",
            json={"key": payload.key, "title": payload.title, "order_index": payload.order_index},
            headers=headers
        )
        if r.status_code not in [200, 201]:
            logger.error(f"Failed to create post: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to create position.")
        return {"success": True}

@elections_router.delete("/posts/{key}")
async def delete_post(key: str):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.delete(
            f"{SUPABASE_URL}/rest/v1/election_posts?key=eq.{key}",
            headers=headers
        )
        if r.status_code not in [200, 204]:
            logger.error(f"Failed to delete post: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to delete position.")
        return {"success": True}

@elections_router.post("/candidates")
async def create_candidate(payload: CandidateCreatePayload):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.post(
            f"{SUPABASE_URL}/rest/v1/election_candidates",
            json={
                "name": payload.name,
                "post_key": payload.post_key,
                "symbol": payload.symbol,
                "photo": payload.photo,
                "symbol_image": payload.symbol_image
            },
            headers=headers
        )
        if r.status_code not in [200, 201]:
            logger.error(f"Failed to nominate candidate: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to nominate candidate.")
        res_data = r.json()
        return {"success": True, "candidate": res_data[0] if res_data else None}

@elections_router.delete("/candidates/{cid}")
async def delete_candidate(cid: str):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.delete(
            f"{SUPABASE_URL}/rest/v1/election_candidates?id=eq.{cid}",
            headers=headers
        )
        if r.status_code not in [200, 204]:
            logger.error(f"Failed to delete candidate: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to delete candidate.")
        return {"success": True}

@elections_router.post("/voters/upload")
async def upload_voters(payload: List[Dict[str, Any]]):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        # 1. Clear existing voters
        r_del = await client.delete(f"{SUPABASE_URL}/rest/v1/election_voters", headers=headers)
        if r_del.status_code not in [200, 204]:
            logger.error(f"Failed to clear voters: {r_del.text}")
            raise HTTPException(status_code=500, detail="Failed to prepare voters database.")

        # 2. Insert new voters in chunks
        for i in range(0, len(payload), 100):
            chunk = payload[i:i + 100]
            r_ins = await client.post(f"{SUPABASE_URL}/rest/v1/election_voters", json=chunk, headers=headers)
            if r_ins.status_code not in [200, 201, 204]:
                logger.error(f"Failed to insert chunk: {r_ins.text}")
                raise HTTPException(status_code=500, detail="Failed to save voter records.")

        return {"success": True, "count": len(payload)}

@elections_router.post("/archive")
async def archive_results(session_name: str = "2026-27"):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        # 1. Fetch posts, candidates, and votes from Supabase
        r_posts = await client.get(f"{SUPABASE_URL}/rest/v1/election_posts", headers=headers)
        r_cands = await client.get(f"{SUPABASE_URL}/rest/v1/election_candidates", headers=headers)
        r_votes = await client.get(f"{SUPABASE_URL}/rest/v1/election_votes", headers=headers)
        
        if r_posts.status_code != 200 or r_cands.status_code != 200 or r_votes.status_code != 200:
            logger.error(f"Failed to fetch archive source tables: posts={r_posts.text}, candidates={r_cands.text}, votes={r_votes.text}")
            raise HTTPException(status_code=500, detail="Failed to fetch data for compilation.")
        
        posts = r_posts.json()
        candidates = r_cands.json()
        votes = r_votes.json()
        
        # Compile counts
        count = {}  # post_key -> candidate_id -> count
        for v in votes:
            sel = v.get("selections") or {}
            for pk, cid in sel.items():
                if pk not in count:
                    count[pk] = {}
                count[pk][cid] = count[pk].get(cid, 0) + 1
                
        archive_rows = []
        for post in posts:
            p_key = post["key"]
            p_title = post["title"]
            post_cands = [c for c in candidates if c["post_key"] == p_key]
            
            post_counts = count.get(p_key, {})
            max_votes = -1
            winner_id = None
            for c in post_cands:
                votes_for_cand = post_counts.get(c["id"], 0)
                if votes_for_cand > max_votes:
                    max_votes = votes_for_cand
                    winner_id = c["id"]
                    
            for c in post_cands:
                v_count = post_counts.get(c["id"], 0)
                archive_rows.append({
                    "session_name": session_name,
                    "post_key": p_key,
                    "post_title": p_title,
                    "candidate_name": c["name"],
                    "candidate_symbol": c["symbol"],
                    "votes_count": v_count,
                    "is_winner": c["id"] == winner_id and v_count > 0
                })
                
        # 2. Insert into archive
        if archive_rows:
            r_ins = await client.post(f"{SUPABASE_URL}/rest/v1/election_results_archive", json=archive_rows, headers=headers)
            if r_ins.status_code not in [200, 201]:
                logger.error(f"Failed to write archive rows: {r_ins.text}")
                raise HTTPException(status_code=500, detail="Failed to save archive results.")
                
        # 3. Clear votes
        r_del_votes = await client.delete(f"{SUPABASE_URL}/rest/v1/election_votes", headers=headers)
        if r_del_votes.status_code not in [200, 204]:
            logger.error(f"Failed to delete votes: {r_del_votes.text}")
            
        # 4. Reset voter flag
        r_reset = await client.patch(
            f"{SUPABASE_URL}/rest/v1/election_voters?already_voted=eq.true",
            json={"already_voted": False},
            headers=headers
        )
        if r_reset.status_code not in [200, 204]:
            logger.error(f"Failed to reset voter flag: {r_reset.text}")
            
        # 5. Set settings to close
        await client.post(
            f"{SUPABASE_URL}/rest/v1/election_settings",
            json={"key": "election_open", "value": "false"},
            headers={**headers, "Prefer": "resolution=merge-duplicates"}
        )
        
        return {"success": True}

from auth import get_current_admin

@elections_router.get("/board")
async def public_board():
    await check_supabase()
    async with httpx.AsyncClient() as client:
        # Fetch posts, voters, and settings
        r_posts = await client.get(f"{SUPABASE_URL}/rest/v1/election_posts", headers=headers)
        r_voters = await client.get(f"{SUPABASE_URL}/rest/v1/election_voters", headers=headers)
        r_votes = await client.get(f"{SUPABASE_URL}/rest/v1/election_votes", headers=headers)
        r_open = await client.get(f"{SUPABASE_URL}/rest/v1/election_settings?key=eq.election_open", headers=headers)

        if r_posts.status_code != 200 or r_voters.status_code != 200 or r_votes.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to load dashboard data.")

        posts = r_posts.json()
        voters = r_voters.json()
        votes = r_votes.json()
        
        election_open = r_open.json()[0]["value"] == "true" if r_open.json() else False

        total_users = len(voters)
        total_students = sum(1 for v in voters if v.get("role") == "student")
        total_teachers = sum(1 for v in voters if v.get("role") == "teacher")

        voted_students = sum(1 for v in voters if v["already_voted"] and v.get("role") == "student")
        voted_teachers = sum(1 for v in voters if v["already_voted"] and v.get("role") == "teacher")

        class_groups = {}
        for u in voters:
            if u.get("role") == "student":
                cls = u.get("class_name") or "Unassigned"
                g = class_groups.setdefault(cls, {"class_name": cls, "total": 0, "voted": 0})
                g["total"] += 1
                if u["already_voted"]:
                    g["voted"] += 1
        class_breakdown = sorted(class_groups.values(), key=lambda x: x["class_name"])

        return {
            "election_open": election_open,
            "categories_count": len(posts),
            "total_users": total_users,
            "total_students": total_students,
            "total_teachers": total_teachers,
            "total_voted": len(votes),
            "voted_students": voted_students,
            "voted_teachers": voted_teachers,
            "pending": max(0, total_users - len(votes)),
            "turnout_pct": round((len(votes) / total_users * 100), 1) if total_users else 0,
            "class_breakdown": class_breakdown,
            "last_vote_at": "",
            "updated_at": ""
        }

@elections_router.get("/results")
async def get_results(admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r_posts = await client.get(f"{SUPABASE_URL}/rest/v1/election_posts?order=order_index.asc", headers=headers)
        r_cands = await client.get(f"{SUPABASE_URL}/rest/v1/election_candidates", headers=headers)
        r_votes = await client.get(f"{SUPABASE_URL}/rest/v1/election_votes", headers=headers)
        r_voters = await client.get(f"{SUPABASE_URL}/rest/v1/election_voters", headers={**headers, "Prefer": "count=exact", "Range": "0-0"})

        if r_posts.status_code != 200 or r_cands.status_code != 200 or r_votes.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to compile results.")

        posts = r_posts.json()
        candidates = r_cands.json()
        votes = r_votes.json()
        
        count_header = r_voters.headers.get("content-range")
        total_users = int(count_header.split("/")[-1]) if count_header and "/" in count_header else 0

        counts = {}
        for v in votes:
            sel = v.get("selections") or {}
            for cid in sel.values():
                counts[cid] = counts.get(cid, 0) + 1

        by_post = {p["key"]: [] for p in posts}
        for c in candidates:
            adj = int(c.get("adjustment") or 0)
            entry = {
                "candidate_id": c["id"],
                "name": c["name"],
                "photo": c.get("photo", ""),
                "symbol": c.get("symbol", ""),
                "votes": counts.get(c["id"], 0) + adj
            }
            if c["post_key"] in by_post:
                by_post[c["post_key"]].append(entry)

        for k in by_post:
            by_post[k].sort(key=lambda x: x["votes"], reverse=True)

        return {
            "posts": [{"key": p["key"], "title": p["title"], "order": p["order_index"]} for p in posts],
            "by_post": by_post,
            "winners": {p["key"]: (by_post[p["key"]][0] if by_post[p["key"]] else None) for p in posts},
            "total_voted": len(votes),
            "total_users": total_users,
            "turnout_pct": round((len(votes) / total_users * 100), 1) if total_users else 0
        }


@elections_router.post("/settings/results_publish_time")
async def set_results_publish_time(payload: Dict[str, Any], admin = Depends(get_current_admin)):
    """Schedule when results become public. payload = {value: ISO datetime string or empty string to clear}"""
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.post(
            f"{SUPABASE_URL}/rest/v1/election_settings",
            json={"key": "results_publish_time", "value": payload.get("value", "")},
            headers={**headers, "Prefer": "resolution=merge-duplicates"}
        )
        if r.status_code not in [200, 201, 204]:
            logger.error(f"Failed to set results_publish_time: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to save schedule.")
        return {"success": True}


@elections_router.get("/public-results")
async def public_results():
    """Public endpoint: returns results only if publish time has passed, otherwise returns countdown."""
    from datetime import datetime, timezone
    import asyncio
    import json
    await check_supabase()
    async with httpx.AsyncClient() as client:
        # 1. Check publish time and appointed post keys settings
        r_settings = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_settings",
            headers=headers
        )
        publish_time_str = ""
        appointed_post_keys = []
        if r_settings.status_code == 200 and r_settings.json():
            for s in r_settings.json():
                if s.get("key") == "results_publish_time":
                    publish_time_str = s.get("value", "")
                elif s.get("key") == "appointed_post_keys":
                    try:
                        appointed_post_keys = json.loads(s.get("value", "[]"))
                    except Exception:
                        pass

        if not publish_time_str:
            return {"status": "sealed", "message": "Results have not been scheduled for release yet."}

        try:
            publish_dt = datetime.fromisoformat(publish_time_str.replace("Z", "+00:00"))
        except Exception:
            return {"status": "sealed", "message": "Invalid publish time configured."}

        now = datetime.now(timezone.utc)
        if now < publish_dt:
            remaining = (publish_dt - now).total_seconds()
            return {
                "status": "countdown",
                "publish_at": publish_time_str,
                "remaining_seconds": max(0, int(remaining)),
                "message": "Results will be published soon."
            }

        # 2. Time has passed — return full results (no auth required)
        # Fetch posts, candidates, votes, and archive concurrently
        tasks = [
            client.get(f"{SUPABASE_URL}/rest/v1/election_posts?order=order_index.asc", headers=headers),
            client.get(f"{SUPABASE_URL}/rest/v1/election_candidates", headers=headers),
            client.get(f"{SUPABASE_URL}/rest/v1/election_votes", headers=headers),
            client.get(f"{SUPABASE_URL}/rest/v1/election_results_archive?order=id.desc&limit=100", headers=headers)
        ]
        r_posts, r_cands, r_votes, r_archive = await asyncio.gather(*tasks)

        if r_posts.status_code != 200 or r_cands.status_code != 200 or r_votes.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to compile results.")

        posts = r_posts.json()
        candidates = r_cands.json()
        votes = r_votes.json()
        archive = r_archive.json() if r_archive.status_code == 200 else []

        # If there are live votes, compute from votes
        if votes:
            counts = {}
            for v in votes:
                sel = v.get("selections") or {}
                for cid in sel.values():
                    counts[cid] = counts.get(cid, 0) + 1

            by_post = {p["key"]: [] for p in posts}
            for c in candidates:
                adj = int(c.get("adjustment") or 0)
                entry = {
                    "candidate_id": c["id"],
                    "name": c["name"],
                    "photo": c.get("photo", ""),
                    "symbol": c.get("symbol", ""),
                    "votes": counts.get(c["id"], 0) + adj
                }
                if c["post_key"] in by_post:
                    by_post[c["post_key"]].append(entry)

            for k in by_post:
                by_post[k].sort(key=lambda x: x["votes"], reverse=True)

            return {
                "status": "live",
                "posts": [{"key": p["key"], "title": p["title"], "order": p["order_index"]} for p in posts],
                "by_post": by_post,
                "winners": {p["key"]: (by_post[p["key"]][0] if by_post[p["key"]] else None) for p in posts},
                "total_voted": len(votes),
                "appointed_post_keys": appointed_post_keys,
            }

        return {"status": "sealed", "message": "No results available."}


# ─────────────────────────────────────────────────────────────────────────────
# EXACT SAME CONTROL PANEL ADMIN ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@elections_router.get("/admin/stats")
async def get_admin_stats(admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        import asyncio
        tasks = [
            client.get(f"{SUPABASE_URL}/rest/v1/election_posts?order=order_index.asc", headers=headers),
            client.get(f"{SUPABASE_URL}/rest/v1/election_candidates", headers=headers),
            client.get(f"{SUPABASE_URL}/rest/v1/election_voters", headers=headers),
            client.get(f"{SUPABASE_URL}/rest/v1/election_votes", headers=headers)
        ]
        r_posts, r_cands, r_voters, r_votes = await asyncio.gather(*tasks)

        if r_posts.status_code != 200 or r_cands.status_code != 200 or r_voters.status_code != 200 or r_votes.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to fetch stats source tables.")

        posts = r_posts.json()
        candidates = r_cands.json()
        voters = r_voters.json()
        votes = r_votes.json()

        total_users = len(voters)
        total_voted = len(votes)
        total_students = sum(1 for v in voters if v.get("role") == "student")
        total_teachers = sum(1 for v in voters if v.get("role") == "teacher")
        voted_students = sum(1 for v in voters if v.get("already_voted") and v.get("role") == "student")
        voted_teachers = sum(1 for v in voters if v.get("already_voted") and v.get("role") == "teacher")
        turnout_pct = round((total_voted / total_users * 100), 1) if total_users else 0

        # Class breakdown
        class_groups = {}
        for u in voters:
            if u.get("role") == "student":
                cls = u.get("class_name") or "Unassigned"
                g = class_groups.setdefault(cls, {"class_name": cls, "total": 0, "voted": 0})
                g["total"] += 1
                if u.get("already_voted"):
                    g["voted"] += 1
        class_breakdown = sorted(class_groups.values(), key=lambda x: x["class_name"])

        # Candidate vote counting
        counts = {}
        for v in votes:
            sel = v.get("selections") or {}
            for cid in sel.values():
                counts[cid] = counts.get(cid, 0) + 1

        by_post = {p["key"]: [] for p in posts}
        for c in candidates:
            adj = int(c.get("adjustment") or 0)
            entry = {
                "candidate_id": c["id"],
                "name": c["name"],
                "photo": c.get("photo", ""),
                "symbol": c.get("symbol", ""),
                "votes": counts.get(c["id"], 0) + adj,
                "real_votes": counts.get(c["id"], 0),
                "adjustment": adj
            }
            if c["post_key"] in by_post:
                by_post[c["post_key"]].append(entry)

        for k in by_post:
            by_post[k].sort(key=lambda x: x["votes"], reverse=True)

        winners = {p["key"]: (by_post[p["key"]][0] if by_post[p["key"]] else None) for p in posts}

        # Voters/votes mapping
        voter_map = {v["admission_no"]: v for v in voters}
        candidate_map = {c["id"]: c for c in candidates}
        votes_list = []
        for vt in votes:
            adm = vt.get("voter_admission_no")
            v_obj = voter_map.get(adm, {})
            sel_names = {}
            for pk, cid in (vt.get("selections") or {}).items():
                sel_names[pk] = candidate_map.get(cid, {}).get("name", "Unknown")
            votes_list.append({
                "id": vt["id"],
                "admission_no": adm,
                "voter_name": v_obj.get("name", "Unknown"),
                "selections": vt.get("selections", {}),
                "selection_names": sel_names
            })

        return {
            "total_users": total_users,
            "total_voted": total_voted,
            "turnout_pct": turnout_pct,
            "total_students": total_students,
            "total_teachers": total_teachers,
            "voted_students": voted_students,
            "voted_teachers": voted_teachers,
            "class_breakdown": class_breakdown,
            "by_post": by_post,
            "winners": winners,
            "votes": votes_list
        }

@elections_router.get("/candidates")
async def get_candidates_list(post: Optional[str] = None):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        url = f"{SUPABASE_URL}/rest/v1/election_candidates"
        if post:
            url += f"?post_key=eq.{post}"
        r = await client.get(url, headers=headers)
        if r.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to fetch candidates.")
        return [{
            "id": c["id"],
            "post": c["post_key"],
            "name": c["name"],
            "photo": c.get("photo", ""),
            "symbol": c.get("symbol", ""),
            "symbol_image": c.get("symbol_image", ""),
            "adjustment": c.get("adjustment", 0)
        } for c in r.json()]

@elections_router.get("/admin/users")
async def list_users(admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.get(f"{SUPABASE_URL}/rest/v1/election_voters?order=admission_no.asc", headers=headers)
        if r.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to fetch voters.")
        
        voters = []
        for v in r.json():
            voters.append({
                "admission_no": v["admission_no"],
                "name": v["name"],
                "role": v["role"],
                "has_voted": v.get("already_voted", False),
                "father_name": v.get("father_name", ""),
                "class_name": v.get("class_name", ""),
                "subject": v.get("subject", ""),
                "designation": v.get("designation", "")
            })
        return voters

@elections_router.delete("/admin/users/{admission_no}")
async def delete_user(admission_no: str, admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.delete(
            f"{SUPABASE_URL}/rest/v1/election_voters?admission_no=eq.{admission_no}",
            headers=headers
        )
        if r.status_code not in [200, 204]:
            raise HTTPException(status_code=500, detail="Failed to delete voter.")
        return {"success": True}

@elections_router.put("/admin/posts/{key}")
async def update_post_admin(key: str, payload: PostCreatePayload, admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.patch(
            f"{SUPABASE_URL}/rest/v1/election_posts?key=eq.{key}",
            json={"title": payload.title, "order_index": payload.order_index},
            headers=headers
        )
        if r.status_code not in [200, 204]:
            logger.error(f"Failed to update post: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to update position.")
        return {"success": True}

@elections_router.delete("/admin/posts/{key}")
async def delete_post_admin(key: str, admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.delete(
            f"{SUPABASE_URL}/rest/v1/election_posts?key=eq.{key}",
            headers=headers
        )
        if r.status_code not in [200, 204]:
            logger.error(f"Failed to delete post: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to delete position.")
        return {"success": True}

@elections_router.post("/admin/posts")
async def create_post_admin(payload: PostCreatePayload, admin = Depends(get_current_admin)):
    return await create_post(payload)

@elections_router.post("/admin/candidates")
async def create_candidate_admin(payload: CandidateCreatePayload, admin = Depends(get_current_admin)):
    return await create_candidate(payload)

class CandidateUpdateAdminPayload(BaseModel):
    name: str
    symbol: str
    photo: Optional[str] = ""
    symbol_image: Optional[str] = ""
    adjustment: Optional[int] = 0
    post_key: Optional[str] = None

@elections_router.put("/admin/candidates/{cid}")
async def update_candidate_admin(cid: str, payload: CandidateUpdateAdminPayload, admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        update_data = {
            "name": payload.name,
            "symbol": payload.symbol,
            "photo": payload.photo,
            "symbol_image": payload.symbol_image,
            "adjustment": payload.adjustment
        }
        if payload.post_key:
            update_data["post_key"] = payload.post_key
            
        r = await client.patch(
            f"{SUPABASE_URL}/rest/v1/election_candidates?id=eq.{cid}",
            json=update_data,
            headers=headers
        )
        if r.status_code not in [200, 204]:
            logger.error(f"Failed to update candidate: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to update candidate.")
        return {"success": True}

@elections_router.delete("/admin/candidates/{cid}")
async def delete_candidate_admin(cid: str, admin = Depends(get_current_admin)):
    return await delete_candidate(cid)

class VoteUpdatePayload(BaseModel):
    selections: Dict[str, str]

@elections_router.put("/admin/votes/{vote_id}")
async def update_vote(vote_id: int, payload: VoteUpdatePayload, admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r = await client.patch(
            f"{SUPABASE_URL}/rest/v1/election_votes?id=eq.{vote_id}",
            json={"selections": payload.selections},
            headers=headers
        )
        if r.status_code not in [200, 204]:
            logger.error(f"Failed to update vote: {r.text}")
            raise HTTPException(status_code=500, detail="Failed to update ballot selections.")
        return {"success": True}

@elections_router.delete("/admin/votes/{vote_id}")
async def delete_vote(vote_id: int, admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        # Find the voter admission no first to reset their flag
        r_vote = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_votes?id=eq.{vote_id}",
            headers=headers
        )
        if r_vote.status_code != 200 or not r_vote.json():
            raise HTTPException(status_code=404, detail="Ballot not found.")
        
        adm = r_vote.json()[0].get("voter_admission_no")
        
        # Delete vote
        r_del = await client.delete(
            f"{SUPABASE_URL}/rest/v1/election_votes?id=eq.{vote_id}",
            headers=headers
        )
        if r_del.status_code not in [200, 204]:
            logger.error(f"Failed to delete vote: {r_del.text}")
            raise HTTPException(status_code=500, detail="Failed to delete ballot.")
        
        # Reset voter flag
        if adm:
            await client.patch(
                f"{SUPABASE_URL}/rest/v1/election_voters?admission_no=eq.{adm}",
                json={"already_voted": False},
                headers=headers
            )
            
        return {"success": True}

@elections_router.get("/admin/settings")
async def get_settings_admin(admin = Depends(get_current_admin)):
    return await get_settings()

@elections_router.put("/admin/settings/{key}")
async def update_settings_admin(key: str, payload: SettingUpdatePayload, admin = Depends(get_current_admin)):
    return await update_settings(key, payload)

@elections_router.post("/admin/reset/votes")
async def reset_votes_only(admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        # Get count of votes first
        r_votes = await client.get(f"{SUPABASE_URL}/rest/v1/election_votes", headers={**headers, "Prefer": "count=exact", "Range": "0-0"})
        ch = r_votes.headers.get("content-range")
        deleted_count = int(ch.split("/")[-1]) if ch and "/" in ch else 0

        # Delete all votes
        r_del = await client.delete(f"{SUPABASE_URL}/rest/v1/election_votes", headers=headers)
        if r_del.status_code not in [200, 204]:
            raise HTTPException(status_code=500, detail="Failed to clear votes.")

        # Reset voters flag
        await client.patch(
            f"{SUPABASE_URL}/rest/v1/election_voters?already_voted=eq.true",
            json={"already_voted": False},
            headers=headers
        )
        return {"success": True, "deleted_votes": deleted_count}

@elections_router.post("/admin/reset/all")
async def reset_everything(admin = Depends(get_current_admin)):
    await check_supabase()
    async with httpx.AsyncClient() as client:
        r_votes = await client.get(f"{SUPABASE_URL}/rest/v1/election_votes", headers={**headers, "Prefer": "count=exact", "Range": "0-0"})
        ch_votes = r_votes.headers.get("content-range")
        del_votes = int(ch_votes.split("/")[-1]) if ch_votes and "/" in ch_votes else 0

        r_cands = await client.get(f"{SUPABASE_URL}/rest/v1/election_candidates", headers={**headers, "Prefer": "count=exact", "Range": "0-0"})
        ch_cands = r_cands.headers.get("content-range")
        del_cands = int(ch_cands.split("/")[-1]) if ch_cands and "/" in ch_cands else 0

        r_voters = await client.get(f"{SUPABASE_URL}/rest/v1/election_voters", headers={**headers, "Prefer": "count=exact", "Range": "0-0"})
        ch_voters = r_voters.headers.get("content-range")
        del_voters = int(ch_voters.split("/")[-1]) if ch_voters and "/" in ch_voters else 0

        await client.delete(f"{SUPABASE_URL}/rest/v1/election_votes", headers=headers)
        await client.delete(f"{SUPABASE_URL}/rest/v1/election_candidates", headers=headers)
        await client.delete(f"{SUPABASE_URL}/rest/v1/election_voters", headers=headers)

        return {
            "success": True,
            "deleted_votes": del_votes,
            "deleted_candidates": del_cands,
            "deleted_users": del_voters
        }

@elections_router.get("/admin/template/{role}")
async def get_excel_template(role: str, admin = Depends(get_current_admin)):
    import io
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    
    if role == "student":
        ws.append(["admission_no", "name", "father_name", "class_name"])
        ws.append(["SDPSS001", "John Doe", "Richard Doe", "Class XII-A"])
    else:
        ws.append(["admission_no", "name", "subject", "designation"])
        ws.append(["SDPSE001", "Jane Smith", "Mathematics", "Senior PGT"])
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sdps_{role}_template.xlsx"}
    )

@elections_router.post("/admin/users/upload")
async def upload_users_admin(role: str, file: UploadFile = File(...), admin = Depends(get_current_admin)):
    import io
    import openpyxl
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    headers_row = [cell.value for cell in ws[1]]
    
    required = ["admission_no", "name"]
    for req in required:
        if req not in headers_row:
            raise HTTPException(status_code=400, detail=f"Missing required column: {req}")
            
    voters = []
    for r_idx in range(2, ws.max_row + 1):
        row_vals = [cell.value for cell in ws[r_idx]]
        if not any(row_vals):
            continue
            
        row_dict = dict(zip(headers_row, row_vals))
        
        adm = str(row_dict.get("admission_no") or "").strip()
        name = str(row_dict.get("name") or "").strip()
        
        if not adm or not name:
            continue
            
        voters.append({
            "admission_no": adm,
            "name": name,
            "role": role,
            "father_name": str(row_dict.get("father_name") or "").strip() if "father_name" in row_dict else None,
            "class_name": str(row_dict.get("class_name") or "").strip() if "class_name" in row_dict else None,
            "subject": str(row_dict.get("subject") or "").strip() if "subject" in row_dict else None,
            "designation": str(row_dict.get("designation") or "").strip() if "designation" in row_dict else None,
            "already_voted": False
        })
        
    if not voters:
        raise HTTPException(status_code=400, detail="No valid voter records found.")
        
    await check_supabase()
    async with httpx.AsyncClient() as client:
        inserted = 0
        updated = 0
        
        r_existing = await client.get(
            f"{SUPABASE_URL}/rest/v1/election_voters?select=admission_no",
            headers=headers
        )
        existing_adms = {x["admission_no"] for x in r_existing.json()} if r_existing.status_code == 200 else set()
        
        for i in range(0, len(voters), 100):
            chunk = voters[i:i+100]
            r_ins = await client.post(
                f"{SUPABASE_URL}/rest/v1/election_voters",
                json=chunk,
                headers={**headers, "Prefer": "resolution=merge-duplicates"}
            )
            if r_ins.status_code not in [200, 201, 204]:
                logger.error(f"Bulk insert error: {r_ins.text}")
                raise HTTPException(status_code=500, detail="Failed to upload voter roster.")
                
            for v in chunk:
                if v["admission_no"] in existing_adms:
                    updated += 1
                else:
                    inserted += 1
                    
        return {"inserted": inserted, "updated": updated}


